from openpyxl import load_workbook
#import textract
import docx2txt
import os

# Directory for files
file_dir = "."
# Caracter separador de colunas
delimitador = "\n-----------------------------------\n"
filename = 'word-to-txt.txt'

with open(file_name, 'w') as fileout:

    # Get all word files
    for f in os.listdir(file_dir):
        if f.endswith(".docx"):
            path = os.path.join(file_dir, f)
            print("Parsing file: {}".format(path))
            # Open the word
            #text = textract.process(path)
            # extract text
            text = docx2txt.process(path)
            # Get file name without extension
            ff = os.path.splitext(f)[0]
            # write to file
            fileout.write( ff + delimitador + str(text) + delimitador )
            '''
            book = load_workbook(path, read_only=True)
            print("Workbook loaded")
            sheets = book.sheetnames
            print(str(len(sheets)) + " Sheets da planilhas: " + str(sheets))
            '''
            # Get file name without extension
            #ff = os.path.splitext(f)[0]
            # Output file generated
            #file_name = "{0}.txt".format(ff)
            print("Writed to file: {}".format(file_name))
            #nlp_file = codecs.open(file_name, 'w', encoding='utf-8')
            '''
            with open(file_name, 'w') as fileout:
                # Iterate through worksheets and get values
                for sheet in book:
                    print("Processando aba {0} ".format(sheet))
                    #print("Processando aba {0} com Dimensões = {1}".format(sheet, sheet.dimensions))
                    #print("Minimum row: {0}".format(sheet.min_row))
                    #print("Maximum row: {0}".format(sheet.max_row))
                    #print("Minimum column: {0}".format(sheet.min_column))
                    #print("Maximum column: {0}".format(sheet.max_column))
                    
                    # Get the  values : nao escreve celulas vazias
                    for linha in range(1, sheet.max_row+1):
                        fileout.write( "\n")
                        for coluna in range(1, sheet.max_column+1):
                            cell = sheet.cell(column=coluna, row=linha).value
                            if cell is not None:
                                fileout.write( str(cell) + delimitador )
                        #Escrevendo a ultima coluna e quebrando a linha
                        #cell = sheet.cell(column=sheet.max_column, row=linha).value
                        #if cell is not None:
                            #fileout.write( str(cell) + "\n")
                    header = sheet.row(0)
                    for rx in range(1, sheet.nrows):
                        for cx in range(0, sheet.ncols):
                            headerVal = convtostr(header[cx].value, 0)
                            if 'date' in headerVal.lower():
                                cellVal = convtostr(sheet.cell_value(rx, cx), 1)
                            else:
                                cellVal = convtostr(sheet.cell_value(rx, cx), 0)
                            if cellVal !="":
                                nlp_text += "{0} is {1} and ".format(headerVal, cellVal)
                                nlp = nlp_text.replace('\n', ' ')
                        nlp += "\n"
            '''

fileout.close()

#book.save('dimensions.xlsx')
